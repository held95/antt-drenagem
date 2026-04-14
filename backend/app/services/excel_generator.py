"""Generate a consolidated Excel file from extracted drainage records."""

from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domain.drainage_record import DrainageRecord

# ---------------------------------------------------------------------------
# Style constants (matching the reference screenshot)
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=11)
DATA_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column definitions: (sub_header, width, group)
COLUMNS: List[Tuple[str, int, str]] = [
    ("Estaca", 12, "Localização do Início"),
    ("Km", 12, "Localização do Início"),
    ("Início Coordenada X", 18, "Localização do Início"),
    ("Início Coordenada Y", 18, "Localização do Início"),
    ("Estaca", 12, "Localização do Fim"),
    ("Km", 12, "Localização do Fim"),
    ("Fim Coordenada X", 18, "Localização do Fim"),
    ("Fim Coordenada Y", 18, "Localização do Fim"),
    ("Altura", 10, "Dimensões"),
    ("Extensão", 12, "Dimensões"),
    ("Largura", 10, "Dimensões"),
    ("Tipo", 12, ""),
    ("Estado de Conservação", 20, ""),
    ("Ambiente", 12, ""),
]


def _apply_cell_style(ws, row: int, col: int, value, font=DATA_FONT,
                       fill: Optional[PatternFill] = None,
                       alignment: Alignment = CENTER,
                       border: Border = THIN_BORDER,
                       number_format: Optional[str] = None):
    """Set value and style for a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    return cell


def _parse_km_sort_key(km: Optional[str]) -> float:
    """Convert km marker like '156+080' to a sortable float (156.080)."""
    if not km:
        return float("inf")
    match = re.match(r"(\d+)\+(\d+)", km)
    if match:
        return int(match.group(1)) + int(match.group(2)) / 1000.0
    return float("inf")


def generate_excel(
    records: Sequence[DrainageRecord],
    highway_name: str = "BR-381/MG Trecho Entr° BR-116/MG (Governador Valadares) - Entr° Belo Horizonte",
) -> bytes:
    """Generate a formatted .xlsx file from drainage records.

    Records are sorted by km_inicial. Returns the file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Drenagem Superficial"

    num_cols = len(COLUMNS)

    # Sort records by km_inicial
    sorted_records = sorted(records, key=lambda r: _parse_km_sort_key(r.km_inicial))

    # -- Row 1: Title --
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value="Rodovia: {}".format(highway_name))
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=1, column=col_idx).border = THIN_BORDER

    # -- Row 2: Group headers --
    groups: List[Tuple[str, int, int]] = []
    current_group = COLUMNS[0][2]
    group_start = 1
    for i, (_, _, group) in enumerate(COLUMNS):
        if group != current_group:
            if current_group:
                groups.append((current_group, group_start, i))
            current_group = group
            group_start = i + 1
        elif i == len(COLUMNS) - 1:
            if current_group:
                groups.append((current_group, group_start, i + 1))

    for group_name, start_col, end_col in groups:
        if start_col < end_col:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        _apply_cell_style(ws, 2, start_col, group_name, font=HEADER_FONT, fill=HEADER_FILL)

    # Fill ungrouped header cells in row 2
    for col_idx, (sub_header, _, group) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL
        if not group:
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
            cell.value = sub_header
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    # -- Row 3: Sub-headers --
    for col_idx, (sub_header, width, group) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if group:
            _apply_cell_style(ws, 3, col_idx, sub_header, font=HEADER_FONT, fill=HEADER_FILL)
        else:
            ws.cell(row=3, column=col_idx).border = THIN_BORDER

    # -- Row 4+: Data rows --
    for row_offset, record in enumerate(sorted_records):
        row_num = 4 + row_offset
        row_data = [
            record.estaca_inicio,
            record.km_inicial,
            record.latitude_inicio,   # Início Coordenada X (~-18.xxx)
            record.longitude_inicio,  # Início Coordenada Y (~-42.xxx)
            record.estaca_fim,
            record.km_final,
            record.latitude_fim,      # Fim Coordenada X (~-18.xxx)
            record.longitude_fim,     # Fim Coordenada Y (~-42.xxx)
            record.altura,
            record.extensao,
            record.largura,
            record.tipo,
            record.estado_conservacao,
            record.ambiente,
        ]
        for col_idx, value in enumerate(row_data, 1):
            fmt = None
            if isinstance(value, float):
                if col_idx in (3, 4, 7, 8):  # Latitude/Longitude columns
                    fmt = "0.000000"
                else:
                    fmt = "0.00"
            _apply_cell_style(ws, row_num, col_idx, value, number_format=fmt)

    # -- Metadata sheet --
    ws_meta = wb.create_sheet("Metadados")
    ws_meta.column_dimensions["A"].width = 30
    ws_meta.column_dimensions["B"].width = 50

    meta_data = [
        ("Data de geração", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Total de PDFs processados", len(sorted_records)),
        ("Rodovia", highway_name),
    ]

    # Count warnings
    total_warnings = sum(len(r.warnings) for r in sorted_records)
    meta_data.append(("Total de avisos", total_warnings))

    low_confidence = [r for r in sorted_records if r.confidence < 0.83]
    if low_confidence:
        meta_data.append(("PDFs com baixa confiança", len(low_confidence)))
        for r in low_confidence:
            meta_data.append((
                "  → {}".format(r.source_filename),
                "confiança: {:.0%} | {}".format(r.confidence, "; ".join(r.warnings) if r.warnings else "OK"),
            ))

    for row_idx, (label, value) in enumerate(meta_data, 1):
        _apply_cell_style(ws_meta, row_idx, 1, label, font=HEADER_FONT, alignment=Alignment(horizontal="left"))
        _apply_cell_style(ws_meta, row_idx, 2, value, alignment=Alignment(horizontal="left"))

    # Write to bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
