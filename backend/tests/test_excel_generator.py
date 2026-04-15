"""Tests for excel_generator module."""

from io import BytesIO

from openpyxl import load_workbook

from app.domain.drainage_record import DrainageRecord
from app.services.excel_generator import generate_excel


def _make_record(**kwargs) -> DrainageRecord:
    """Create a DrainageRecord with defaults for required fields."""
    defaults = dict(
        source_filename="test.pdf",
        estaca_inicio="156+080",
        km_inicial="156+080",
        latitude_inicio=-18.91475,
        longitude_inicio=-42.61853,
        estaca_fim="156+390",
        km_final="156+390",
        latitude_fim=-18.91086,
        longitude_fim=-42.0125,
        largura=0.15,
        altura=0.2,
        extensao=180.0,
        tipo="MFC/VG",
        estado_conservacao="REGULAR",
        ambiente="URBANO",
        confidence=1.0,
    )
    defaults.update(kwargs)
    return DrainageRecord(**defaults)


def _load_wb(excel_bytes: bytes):
    return load_workbook(BytesIO(excel_bytes))


class TestGenerateExcel:
    def test_returns_bytes(self):
        records = [_make_record()]
        result = generate_excel(records)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_title_row(self):
        wb = _load_wb(generate_excel([_make_record()]))
        ws = wb.active
        assert ws.cell(1, 1).value.startswith("Rodovia:")

    def test_group_headers_merged(self):
        wb = _load_wb(generate_excel([_make_record()]))
        ws = wb.active
        merged = [str(m) for m in ws.merged_cells.ranges]
        # Title row should be merged
        assert any("A1" in m for m in merged)
        # Group headers should be merged
        assert any("A2" in m for m in merged)

    def test_data_starts_row_4(self):
        wb = _load_wb(generate_excel([_make_record()]))
        ws = wb.active
        assert ws.cell(4, 3).value == "156+080"  # estaca_inicio (col 3, after Data Insp. and Identificação)

    def test_coordinate_format(self):
        wb = _load_wb(generate_excel([_make_record()]))
        ws = wb.active
        cell = ws.cell(4, 5)  # latitude_inicio column (col 5, after Data Insp., Identificação, Estaca, Km)
        assert cell.number_format == "0.000000"

    def test_dimension_format(self):
        wb = _load_wb(generate_excel([_make_record()]))
        ws = wb.active
        cell = ws.cell(4, 11)  # Altura column (col 11, after 2 id + 4 início + 4 fim)
        assert cell.number_format == "0.00"

    def test_multiple_records(self):
        records = [
            _make_record(source_filename="a.pdf", km_inicial="200+000"),
            _make_record(source_filename="b.pdf", km_inicial="100+000"),
            _make_record(source_filename="c.pdf", km_inicial="150+000"),
        ]
        wb = _load_wb(generate_excel(records))
        ws = wb.active
        # Should be sorted by km_inicial — Km (Início) is now col 4
        assert ws.cell(4, 4).value == "100+000"
        assert ws.cell(5, 4).value == "150+000"
        assert ws.cell(6, 4).value == "200+000"

    def test_metadata_sheet(self):
        wb = _load_wb(generate_excel([_make_record()]))
        assert "Metadados" in wb.sheetnames
        ws_meta = wb["Metadados"]
        assert ws_meta.cell(1, 1).value == "Data de geração"
        assert ws_meta.cell(2, 2).value == 1  # total PDFs

    def test_empty_records(self):
        result = generate_excel([])
        wb = _load_wb(result)
        ws = wb.active
        assert ws.cell(1, 1).value.startswith("Rodovia:")
        # No data rows, row 4 should be empty
        assert ws.cell(4, 1).value is None
